import sys
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (PatternFill, colors)
import PySimpleGUI as sg

# program will combine data from 2 CSV files
# rev 0.1a based on GSX completed repair reports and GSX parts pending reports. IT4 reports not used


# GUI to open CSV and XLSX file
sg.theme("DarkTeal2")
layout = [[sg.Text("Виберіть CSV з Completed Repairs")], [sg.T(""), sg.Input(),
                       sg.FileBrowse(key="file1", file_types=(('CSV Files', 'repair*.csv'),))],
          [sg.Text("Виберіть CSV з All Returns:")], [sg.T(""), sg.Input(),
                       sg.FileBrowse(key="file2", file_types=(('CSV Files', 'orderreturn*.csv'),))],
          [sg.Button("Сформувати звіт")]]
window = sg.Window('GSX Report, rev. 0.1a', layout, size=(450, 150))

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == "Exit":
        break
    elif event == "Сформувати звіт" and values["file1"] != '' and values["file2"] != '':
        gsx_comp = open(values["file1"], encoding='Windows-1251')
        gsx_comp_data = csv.DictReader(gsx_comp, delimiter=',')
        gsx = open(values["file2"], encoding='Windows-1251')
        gsx_data = csv.DictReader(gsx, delimiter=',')
        report = Workbook()
        report_sheet = report.create_sheet("Sheet1")
        gsx_list = []
        for rows in gsx_data:
            gsx_list.append(rows)
        counter = 0
        rep_counter = 0
        for i in gsx_comp_data:
            rep_counter += 1
            found_in_gsx = False
            for g in gsx_list:
                if g['Part'] == '011-00213' or g['Part'] == '011-00212' or g['Part'] == '011-00214':
                    continue
                if g['Reference'] == i['Reference']:
                    counter += 1
                    report_sheet.cell(row=counter, column=1, value=g['Part'])  # Part number
                    report_sheet.cell(row=counter, column=2, value=g['Part Description'])  # Part descr
                    report_sheet.cell(row=counter, column=4, value=i['Mark Complete Date'])  # date received
                    report_sheet.cell(row=counter, column=19, value=i['Reference'])
                    report_sheet.cell(row=counter, column=20, value=i['Serial Number'])
                    report_sheet.cell(row=counter, column=22, value=g['Service Notification'])
                    report_sheet.cell(row=counter, column=23, value=g['Repair'])
                    if g['п»ї"Return Order"'] == '':
                        report_sheet.cell(row=counter, column=24, value='NRET')
                    else:
                        report_sheet.cell(row=counter, column=24, value=g['п»ї"Return Order"'])
                    if g['Coverage Status'] == 'Out Of Warranty (No Coverage)':
                        report_sheet.cell(row=counter, column=25, value='Out Of Warranty')
                        report_sheet.cell(row=counter, column=26, value=0.00)
                    else:
                        report_sheet.cell(row=counter, column=25, value='Warranty')
                    if counter > 1 and g['Reference'] == report_sheet.cell(row=counter - 1, column=19).value:
                        report_sheet.cell(row=counter, column=26, value=0.00)
                        for x in range(1, report_sheet.max_column + 1):
                            report_sheet.cell(row=counter, column=x).fill = PatternFill("solid", fgColor="EFF6E3")
                    if g['Serial Number'] != i['Serial Number'] or 'omplete' not in g['Repair Status']:
                        for y in range(1, report_sheet.max_column + 1):
                            report_sheet.cell(row=counter, column=y).fill = PatternFill("solid", fgColor="F44336")
                    found_in_gsx = True
            if found_in_gsx == False:
                counter += 1
                report_sheet.cell(row=counter, column=4, value=i['Mark Complete Date'])  # date received
                report_sheet.cell(row=counter, column=1, value=i['Part'])  # part number
                report_sheet.cell(row=counter, column=2,
                                  value=i['Part Description'])  # part description
                report_sheet.cell(row=counter, column=19, value=i['Reference'])
                report_sheet.cell(row=counter, column=20, value=i['Serial Number'])
                for y in range(1, report_sheet.max_column + 1):
                    report_sheet.cell(row=counter, column=y).fill = PatternFill("solid", fgColor="FFD966")
        report_sheet.cell(row=counter+2, column=1).fill = PatternFill("solid", fgColor="EFF6E3")
        report_sheet.cell(row=counter+2, column=3, value='кольором позначені 2а і наступні запчсатини в ремонті')
        report_sheet.cell(row=counter+4, column=1).fill = PatternFill("solid", fgColor="F44336")
        report_sheet.cell(row=counter+4, column=3, value='потребує уваги, можливі помилки')
        report_sheet.cell(row=counter+6, column=1).fill = PatternFill("solid", fgColor="FFD966")
        report_sheet.cell(row=counter+6, column=3, value='ремонт відсутній в таблиці "All parts"')

        window.close()
# it4.close()
# gsx.close()

# GUI to set destination file folder
layout = [[sg.Text("Виберіть папку для збереження звіту: ")], [sg.T(""), sg.Input(), sg.FolderBrowse(key="folder1")],

          [sg.Text(f"Знайдено ремонтів: {rep_counter}")],
          [sg.T("")],

          [sg.Button("Зберегти звіт")]]

window = sg.Window('GSX Report', layout, size=(450, 150))

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == "Exit":
        break
    elif event == "Зберегти звіт":
        try:
            del report['Sheet']
            report.save(f'{values["folder1"]}/GSXreport.xlsx')
        except:
            sg.popup( f'Виникла помилка при збереженні файлу.\n'
                                          f'Перевірте чи файл GSXreport.xlsx не відкрито.\n'
                                          f'Якщо так, закрийте файл в Excel та запустіть додаток ще раз.\n', title='Помилка')
        window.close()


